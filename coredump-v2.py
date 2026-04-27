#!/usr/bin/env python3
"""
Coredump Auto-Processing Tool v2

Single Python script that orchestrates the full coredump analysis pipeline:
  Step 1: ELK Query
  Step 2: First confirmation (per record)
  Step 3: Download (per SN)
  Step 4: Second confirmation & Jira post
  Step 5: Upload FTP
  Step 6: Merge CSV → XLSX
  Step 7: Report
  Step 8: Cleanup

Usage:
    python3 coredump-v2.py                        # query yesterday
    python3 coredump-v2.py -d 3                   # query last 3 days
    python3 coredump-v2.py --date 2026-03-20      # query specific date
    python3 coredump-v2.py --head                 # show browser
    python3 coredump-v2.py --resume               # resume unfinished work
    python3 coredump-v2.py --retry [SN]           # retry failed downloads
"""

import argparse
import csv
import ftplib
import importlib
import json
import os
import re
import shutil
import sys
import time
import uuid
from datetime import datetime, timedelta, timezone

import openpyxl
import requests
from requests.auth import HTTPBasicAuth
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

from config_loader import load_config
from logger import Logger
from csv_helper import load_csv, save_csv, update_daily_csv, JIRA_ID_COLUMNS
from jira_api import (
    get_custom_field_map, create_issue, update_description, update_fields, check_issue_exists,
    verify_issue_fields,
)

# ─── Import from hyphenated module names ──────────────────────────────────────
_elk_query = importlib.import_module("ELK-query")

# ─── Config ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
STATUS_FILE = os.path.join(SCRIPT_DIR, "status.json")
LOG_DIR = os.path.join(SCRIPT_DIR, "logs")
COREDUMPS_DIR = os.path.join(SCRIPT_DIR, "coredumps")

_config = load_config(extra_keys=[
    "ES-url", "optools-host", "optools-user", "optools-pass",
    "ftp-host", "ftp-user", "ftp-pass", "parent-SF", "sharepoint-url",
])

USER_CONFIRM = _config.get("user-confirm", True)
MAX_LOG_FILES = _config.get("max-log-files", 100)
MAX_COREDUMP_DIRS = _config.get("max-coredump-dirs", 30)

# OpTools
_optools_user = _config["optools-user"]
_optools_pass = _config["optools-pass"]
_optools_host = _config["optools-host"]
_optools_user_encoded = _optools_user.replace("@", "%40") + ":" + _optools_pass
OPTOOLS_BASE_URL = f"https://{_optools_user_encoded}@{_optools_host}"
OPTOOLS_DEVICE_TAB = f"{OPTOOLS_BASE_URL}#tabs-device"
OPTOOLS_QUERY_TAB = f"{OPTOOLS_BASE_URL}#tabs-3"

POLL_INTERVAL = _config.get("poll-interval", 5)
POLL_TIMEOUT = _config.get("poll-timeout", 120)
DOWNLOAD_TIMEOUT = _config.get("download-timeout", 120)
DOWNLOAD_RETRIES = _config.get("download-retries", 3)
DOWNLOAD_RETRY_DELAY = _config.get("download-retry-delay", 5)

TZ_TAIPEI = timezone(timedelta(hours=8))
SKIP_JIRA_VALUES = ("", "Skip", "Fail", "OPTOOLS fail", "OPTOOLS mis-match")

# Jira
PROJECT_KEY = "ZNGA"
ISSUE_TYPE_NAME = "漏洞"


# ═══════════════════════════════════════════════════════════════════════════════
# Utility Functions
# ═══════════════════════════════════════════════════════════════════════════════

def extract_coredump_key(target):
    """Extract coredump-key from target filename.

    Example: 260322-194917-1.37_ABXF.1_-157fc_libedit-nc-cli.core.zip
    → coredump-key: 157fc_libedit-nc-cli.core.zip
    """
    if not target:
        return ""
    filename = os.path.basename(target)
    suffix = filename[14:] if len(filename) > 14 else filename
    idx = suffix.find('-')
    if idx >= 0:
        return suffix[idx + 1:]
    return suffix


def extract_coredump_files(target):
    """從 target 路徑取得檔名，移除前 14 個字元（日期時間）"""
    filename = os.path.basename(target)
    return filename[14:] if len(filename) > 14 else filename


def is_its_firmware(firmware):
    return 'ITS' in firmware.upper() if firmware else False


def parse_affects_version(firmware):
    """Parse firmware version to Jira Affects Version format.

    Examples:
        1.37(ABZH.1) -> 1.37 p1c0
        1.37(ACII.1)b3 -> 1.37 p1b3
        1.37(ABZI.1)Italy -> 1.37 p1c0  (strip "Italy")
        1.37(ABZI.0)ITS-26WK04-m10433 -> 1.37 p0c0  (strip ITS info)
    """
    match = re.match(r'^([\d.]+)\([^.]+\.(\d+)\)(.*)', firmware)
    if not match:
        return firmware

    major, patch, suffix = match.group(1), match.group(2), match.group(3)

    # Clean suffix: only keep valid version characters (b, c, s, digits)
    # Strip out country codes, ITS branches, etc.
    if suffix:
        # Extract valid version suffix like b3, c0, s1
        valid_suffix = re.match(r'^([bcs]\d+)', suffix)
        if valid_suffix:
            return f"{major} p{patch}{valid_suffix.group(1)}"

    # Default to c0 if no valid suffix
    return f"{major} p{patch}c0"


def generate_temp_id():
    return f"tmp-{uuid.uuid4().hex[:8]}"


# ═══════════════════════════════════════════════════════════════════════════════
# Status JSON Management
# ═══════════════════════════════════════════════════════════════════════════════

def load_status():
    if os.path.exists(STATUS_FILE):
        with open(STATUS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"session": {}, "records": []}


def save_status(status):
    tmp = STATUS_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)
    os.replace(tmp, STATUS_FILE)


def update_status_session(status, updates):
    status["session"].update(updates)
    status["session"]["updated"] = datetime.now().isoformat()
    save_status(status)


def has_unfinished_work(status):
    for rec in status.get("records", []):
        if rec.get("status") in ("confirmed", "download_ok", "download_fail"):
            return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# User Input
# ═══════════════════════════════════════════════════════════════════════════════

def ask_user(prompt, valid_choices, default=None):
    if not USER_CONFIRM and default is not None:
        return default
    while True:
        ans = input(f"\n>>> {prompt}: ").strip().lower()
        if ans in valid_choices:
            return ans
        if not ans and default:
            return default
        print(f"請輸入 {'/'.join(valid_choices)}")


# ═══════════════════════════════════════════════════════════════════════════════
# Step 1: ELK Query
# ═══════════════════════════════════════════════════════════════════════════════

def step1_elk_query(days=1, date=None, logger=None):
    logger.log("=" * 60)
    logger.log("Step 1: ELK Query")
    logger.log("=" * 60)

    start_time, end_time = _elk_query.get_date_range(days, specific_date=date)
    logger.log(f"Querying: {start_time} ~ {end_time}")

    records = _elk_query.query_elasticsearch(start_time, end_time, logger=logger)
    if not records:
        logger.log("No matching records found.")
        return None

    if date:
        date_str = date
    else:
        end_day = datetime.now(TZ_TAIPEI) - timedelta(days=1)
        date_str = end_day.strftime("%Y-%m-%d")

    daily_csv = os.path.join(SCRIPT_DIR, f"{date_str}.csv")
    _elk_query.write_csv(records, daily_csv, logger=logger)

    merge_logger = Logger(source="ELK-merge")
    _elk_query.merge_to_summary(daily_csv, _elk_query.SUMMARY_FILE, merge_logger)
    merge_logger.close()

    logger.log(f"ELK Query 完成，{len(records)} 筆記錄")
    return date_str


# ═══════════════════════════════════════════════════════════════════════════════
# Step 2: First Confirmation
# ═══════════════════════════════════════════════════════════════════════════════

def step2_first_confirm(status, logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 2: 第一次確認")
    logger.log("=" * 60)

    csv_path, fieldnames, rows = load_csv()

    # Collect existing coredump-keys with jira-ids (from previous runs)
    existing_keys = {}      # coredump_key -> jira_id (non-ITS)
    its_existing_keys = {}  # coredump_key -> jira_id (ITS)
    for row in rows:
        ck = extract_coredump_key(row.get("target", ""))
        if not ck:
            continue
        jid = (row.get("jira-id") or "").strip()
        its_jid = (row.get("ITS-jira-id") or "").strip()
        if jid and jid not in SKIP_JIRA_VALUES and ck not in existing_keys:
            existing_keys[ck] = jid
        if its_jid and its_jid not in SKIP_JIRA_VALUES and ck not in its_existing_keys:
            its_existing_keys[ck] = its_jid

    # Find pending records (all four jira columns empty)
    pending = []
    for i, row in enumerate(rows):
        jid = (row.get("jira-id") or "").strip()
        rid = (row.get("related-jira-id") or "").strip()
        its_jid = (row.get("ITS-jira-id") or "").strip()
        its_rid = (row.get("ITS-related-jira-id") or "").strip()
        if not (jid or rid or its_jid or its_rid):
            pending.append((i, row))

    if not pending:
        logger.log("沒有待處理的記錄。")
        return status

    # ── 2a: Pre-scan to classify pending records ──
    auto_related = 0    # coredump-key already exists in CSV
    batch_dup = 0       # duplicate coredump-key within this batch
    its_new = 0         # new ITS records
    non_its_new = 0     # new non-ITS records
    seen_keys = set()

    for _, row in pending:
        ck = extract_coredump_key(row.get("target", ""))
        firmware = row.get("firmware", "").strip()
        is_its = is_its_firmware(firmware)

        if is_its:
            existing_jid = its_existing_keys.get(ck)
        else:
            existing_jid = existing_keys.get(ck)

        if existing_jid:
            auto_related += 1
        elif ck in seen_keys:
            batch_dup += 1
        else:
            seen_keys.add(ck)
            if is_its:
                its_new += 1
            else:
                non_its_new += 1

    need_confirm = non_its_new + its_new + batch_dup

    # ── 2b: Summary ──
    logger.log(f"\n--- Summary ---")
    logger.log(f"  待處理總筆數:                 {len(pending)}")
    logger.log(f"  已有重複 key (自動 related):  {auto_related}")
    logger.log(f"  本次 batch 內重複:            {batch_dup}")
    logger.log(f"  新 coredump-key (非 ITS):     {non_its_new}")
    logger.log(f"  新 coredump-key (ITS):        {its_new}")
    logger.log(f"  需要確認:                     {need_confirm}")

    # ── 2c: List all items that need confirmation ──
    if need_confirm > 0:
        logger.log(f"\n--- 待確認清單 ---")
        list_idx = 0
        list_seen_keys = set()
        for _, row in pending:
            ck = extract_coredump_key(row.get("target", ""))
            firmware = row.get("firmware", "").strip()
            is_its = is_its_firmware(firmware)
            sn = row.get("sn", "").strip()
            daemon = row.get("daemon", "").strip()
            model = row.get("model", "").strip()

            if is_its:
                existing_jid = its_existing_keys.get(ck)
            else:
                existing_jid = existing_keys.get(ck)

            if existing_jid:
                continue  # auto related, skip listing

            list_idx += 1
            dup_tag = " [batch 重複]" if ck in list_seen_keys else ""
            its_tag = " [ITS]" if is_its else ""
            list_seen_keys.add(ck)

            logger.log(f"  ({list_idx}/{need_confirm}) SN: {sn}, Daemon: {daemon}, "
                       f"Model: {model}{its_tag}{dup_tag}")
            logger.log(f"    Key: {ck}")

    # ── 2d: Per-record confirmation ──
    logger.log(f"\n--- 逐筆確認 ---")

    # Track batch-level coredump-key assignments
    batch_keys = {}      # coredump_key -> temp_id (non-ITS)
    its_batch_keys = {}  # coredump_key -> temp_id (ITS)

    confirmed_count = 0
    related_count = 0
    skip_count = 0
    ignore_count = 0

    for current, (i, row) in enumerate(pending, 1):
        target = row.get("target", "")
        ck = extract_coredump_key(target)
        firmware = row.get("firmware", "").strip()
        is_its = is_its_firmware(firmware)
        sn = row.get("sn", "").strip()
        daemon = row.get("daemon", "").strip()
        model = row.get("model", "").strip()
        coredump_files = extract_coredump_files(target)

        # ── Case 1: existing in CSV from previous runs → auto related ──
        if is_its:
            existing_jid = its_existing_keys.get(ck)
            target_col_related = "ITS-related-jira-id"
        else:
            existing_jid = existing_keys.get(ck)
            target_col_related = "related-jira-id"

        if existing_jid:
            rows[i][target_col_related] = existing_jid
            save_csv(csv_path, fieldnames, rows)
            update_daily_csv(row.get("_id", ""), target_col_related, existing_jid)
            related_count += 1
            logger.log(f"({current}/{len(pending)}) 已有重複 coredump-key，自動填 {target_col_related}={existing_jid}")
            logger.log(f"  SN: {sn}, Daemon: {daemon}, Key: {ck}")
            continue

        # ── Case 2: duplicate in this batch → ask user ──
        batch_temp = (its_batch_keys if is_its else batch_keys).get(ck)
        if batch_temp:
            logger.log(f"\n{'='*60}")
            logger.log(f"({current}/{len(pending)}) 本次 batch 內重複 coredump-key")
            logger.log(f"  ITS:            {'Yes' if is_its else 'No'}")
            logger.log(f"  SN:             {sn}")
            logger.log(f"  Daemon:         {daemon}")
            logger.log(f"  Model:          {model}")
            logger.log(f"  Firmware:       {firmware}")
            logger.log(f"  Coredump Key:   {ck}")
            logger.log(f"  對應 temp_id:   {batch_temp}")

            default = "s" if is_its else "y"
            choice = ask_user(
                f"({current}/{len(pending)}) 填入 related? (y=填入 / s=跳過)",
                ["y", "s"], default=default,
            )

            if choice == "y":
                target_col = "ITS-related-jira-id" if is_its else "related-jira-id"
                rows[i][target_col] = batch_temp
                save_csv(csv_path, fieldnames, rows)
                update_daily_csv(row.get("_id", ""), target_col, batch_temp)
                related_count += 1

                status["records"].append({
                    "_id": row.get("_id", ""), "sn": sn, "daemon": daemon,
                    "firmware": firmware, "model": model,
                    "coredump_key": ck, "coredump_files": coredump_files,
                    "temp_id": batch_temp, "status": "confirmed",
                    "is_related": True, "is_its": is_its,
                    "jira_id": None, "jira_col": target_col,
                    "download_path": None, "error": None, "retries": 0,
                })
                save_status(status)
                logger.log(f"  已填入 {target_col}={batch_temp}")
            else:
                skip_col = "ITS-jira-id" if is_its else "jira-id"
                rows[i][skip_col] = "Skip"
                save_csv(csv_path, fieldnames, rows)
                update_daily_csv(row.get("_id", ""), skip_col, "Skip")
                skip_count += 1
            continue

        # ── Case 3: new coredump-key → show info and ask ──
        target_col = "ITS-jira-id" if is_its else "jira-id"

        logger.log(f"\n{'='*60}")
        logger.log(f"({current}/{len(pending)})")
        logger.log(f"  ITS:              {'Yes' if is_its else 'No'}")
        logger.log(f"  Firmware:         {firmware}")
        logger.log(f"  影響版本:         {parse_affects_version(firmware)}")
        logger.log(f"  Serial Number:    {sn}")
        logger.log(f"  Model:            {model}")
        logger.log(f"  Daemon:           {daemon}")
        logger.log(f"  Coredump File(s): {coredump_files}")
        logger.log(f"  Coredump Key:     {ck}")
        logger.log(f"  目標欄位:         {target_col}")

        default = "s" if is_its else "y"
        choice = ask_user(
            f"({current}/{len(pending)}) 要處理嗎？(y=確認 / s=跳過 / i=忽略 / q=結束)",
            ["y", "s", "i", "q"], default=default,
        )

        if choice == "q":
            logger.log("使用者選擇結束。")
            break
        elif choice == "s":
            rows[i][target_col] = "Skip"
            save_csv(csv_path, fieldnames, rows)
            update_daily_csv(row.get("_id", ""), target_col, "Skip")
            skip_count += 1
        elif choice == "i":
            ignore_count += 1
        elif choice == "y":
            temp_id = generate_temp_id()
            if is_its:
                its_batch_keys[ck] = temp_id
            else:
                batch_keys[ck] = temp_id

            rows[i][target_col] = temp_id
            save_csv(csv_path, fieldnames, rows)
            update_daily_csv(row.get("_id", ""), target_col, temp_id)

            status["records"].append({
                "_id": row.get("_id", ""), "sn": sn, "daemon": daemon,
                "firmware": firmware, "model": model,
                "coredump_key": ck, "coredump_files": coredump_files,
                "temp_id": temp_id, "status": "confirmed",
                "is_related": False, "is_its": is_its,
                "jira_id": None, "jira_col": target_col,
                "download_path": None, "error": None, "retries": 0,
            })
            save_status(status)
            confirmed_count += 1
            logger.log(f"  已確認，temp_id: {temp_id}")

    update_status_session(status, {"current_step": "confirmed"})

    logger.log(f"\n{'='*60}")
    logger.log(f"第一次確認結果：確認 {confirmed_count}，重複 {related_count}，"
               f"跳過 {skip_count}，忽略 {ignore_count}")
    return status


# ═══════════════════════════════════════════════════════════════════════════════
# Step 3: Download (Playwright)
# ═══════════════════════════════════════════════════════════════════════════════

def _get_mac_from_optools(page, sn, logger):
    """Use OpTools web UI to look up MAC address for a SN."""
    logger.log(f"OpTools: 查詢 SN {sn} 的 MAC address...")
    page.goto(OPTOOLS_DEVICE_TAB, wait_until="networkidle", timeout=60000)
    page.wait_for_timeout(3000)
    page.evaluate("try { $('.loadingoverlay').remove(); } catch(e) {}")
    page.wait_for_timeout(1000)

    sn_inputs = page.locator("#get-device-by-sn-input")
    for idx in range(sn_inputs.count()):
        sn_inputs.nth(idx).fill(sn)

    page.evaluate("""(() => {
        var btn = $('#get-device-by-mac-sn-button');
        if (btn.length === 0) return;
        var events = $._data(btn[0], 'events');
        if (!(events && events.click) && typeof window.initDeviceTab === 'function')
            window.initDeviceTab();
    })()""")

    page.locator("#get-device-by-mac-sn-button").first.click(force=True)

    result_text = ""
    poll_start = time.time()
    while time.time() - poll_start < 30:
        result_text = page.evaluate("$('#get-device-result').text().trim()")
        if result_text:
            break
        page.wait_for_timeout(1000)

    page.evaluate("try { $('.loadingoverlay').remove(); } catch(e) {}")
    page.screenshot(path=os.path.join(LOG_DIR, "v2_mac_result.png"))

    mac_match = re.search(r'([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})', result_text)
    mac = mac_match.group(0) if mac_match else None

    if mac:
        logger.log(f"  MAC found: {mac}")
    else:
        logger.log(f"  [FAIL] MAC not found for SN {sn}")
    return mac


def _download_from_optools(page, sn, mac, logger):
    """Use OpTools to trigger upload and poll for download link."""
    logger.log(f"OpTools: 下載 SN {sn} (MAC: {mac})...")
    page.goto(OPTOOLS_QUERY_TAB, wait_until="networkidle", timeout=60000)
    page.wait_for_timeout(3000)
    page.evaluate("document.querySelectorAll('.loadingoverlay').forEach(el => el.remove());")
    page.wait_for_timeout(1000)

    page.evaluate(f"""(() => {{
        $('#detail-log-mac').val('{mac}');
        $('#detail-log-sn').val('{sn}');
        $('#detail-log-textarea').val('test');
    }})()""")

    verify = page.evaluate("""(() => ({
        mac: $('#detail-log-mac').val(),
        sn: $('#detail-log-sn').val(),
        msg: $('#detail-log-textarea').val()
    }))()""")
    logger.log(f"  Verified: mac={verify['mac']} sn={verify['sn']}")

    page.evaluate("if (!$.LoadingOverlay) { $.LoadingOverlay = function() {}; }")

    last_alert = {"msg": None}

    def on_dialog(dialog):
        last_alert["msg"] = dialog.message
        logger.log(f"  [DIALOG] {dialog.type}: {dialog.message}")
        dialog.accept()

    page.on("dialog", on_dialog)

    def on_request(req):
        if "dev-troubleshoot" in req.url:
            logger.log(f"  [NET] Request: {req.method} {req.url}")

    def on_response(resp):
        if "dev-troubleshoot" in resp.url:
            logger.log(f"  [NET] Response: {resp.status} {resp.url}")

    page.on("request", on_request)
    page.on("response", on_response)

    # Trigger upload
    logger.log("  Calling trigger_device_upload_log()...")
    last_alert["msg"] = None
    page.evaluate("trigger_device_upload_log()")

    start = time.time()
    success = False
    while time.time() - start < 30:
        if last_alert["msg"]:
            if "Succeed" in last_alert["msg"] or "Success" in last_alert["msg"]:
                logger.log("  [OK] Upload succeeded!")
                success = True
            break
        page.wait_for_timeout(500)

    if not success:
        logger.log("  [WARN] Upload may not have succeeded.")

    key_path = page.evaluate("$('#key-path').val() || 'not found'")
    logger.log(f"  Key path: {key_path}")

    # Poll for download link
    logger.log("  Polling for download link...")
    final_link = None
    total_waited = 0
    attempt = 0

    while total_waited < POLL_TIMEOUT:
        attempt += 1
        logger.log(f"  Attempt {attempt}: calling check_device_log()...")
        last_alert["msg"] = None
        page.evaluate("check_device_log()")
        page.wait_for_timeout(POLL_INTERVAL * 1000)
        total_waited += POLL_INTERVAL

        if last_alert["msg"]:
            logger.log(f"    Alert: {last_alert['msg']}")

        check_result = page.evaluate("""(() => {
            const span = document.getElementById('detail-log-check-result');
            const a = document.getElementById('detail-log-check-result-a');
            return {
                text: span ? span.innerText : '',
                href: a ? (a.getAttribute('href') || '') : '',
                linkText: a ? a.innerText : ''
            };
        })()""")

        href = check_result.get("href", "")
        if href and href.strip() and not href.endswith("#") and "index.html" not in href:
            final_link = href.strip()
            logger.log(f"  [OK] Download link found: {final_link}")
            break

        logger.log(f"    Waited {total_waited}s / {POLL_TIMEOUT}s")

    page.screenshot(path=os.path.join(LOG_DIR, "v2_download_result.png"))

    try:
        page.remove_listener("dialog", on_dialog)
        page.remove_listener("request", on_request)
        page.remove_listener("response", on_response)
    except Exception:
        pass

    return final_link


def _verify_coredump(filepath, coredump_key, logger):
    """Verify downloaded zip contains the expected coredump-key.

    Runs 'unzip -l' and checks if any entry contains the coredump-key.
    Returns (ok, zip_listing) tuple.
    """
    import subprocess
    logger.log(f"  Verifying zip contents for key: {coredump_key}")
    try:
        result = subprocess.run(
            ["unzip", "-l", filepath],
            capture_output=True, text=True, timeout=30,
        )
        listing = result.stdout
        logger.log(f"  unzip -l output:\n{listing}")

        if coredump_key and coredump_key in listing:
            logger.log(f"  [OK] Coredump key found in zip")
            return True, listing
        else:
            logger.log(f"  [FAIL] Coredump key NOT found in zip")
            return False, listing
    except Exception as e:
        logger.log(f"  [WARN] unzip verification failed: {e}")
        return False, str(e)


def _download_file(url, dest_dir, filename, logger):
    """Download file from URL with retry."""
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, filename)
    logger.log(f"  Downloading: {url}")
    logger.log(f"  Saving to: {dest}")

    last_error = None
    for attempt in range(1, DOWNLOAD_RETRIES + 1):
        try:
            resp = requests.get(
                url, verify=False, timeout=DOWNLOAD_TIMEOUT, stream=True,
                auth=(_optools_user, _optools_pass),
            )
            resp.raise_for_status()
            total = 0
            with open(dest, "wb") as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)
                    total += len(chunk)
            logger.log(f"  [OK] Downloaded {total / 1024:.1f} KB -> {dest}")
            return dest
        except (requests.RequestException, IOError) as e:
            last_error = e
            if attempt < DOWNLOAD_RETRIES:
                logger.log(f"  [WARN] Attempt {attempt}/{DOWNLOAD_RETRIES} failed: {e}, "
                           f"retrying in {DOWNLOAD_RETRY_DELAY}s...")
                time.sleep(DOWNLOAD_RETRY_DELAY)
            else:
                logger.log(f"  [ERROR] Download failed after {DOWNLOAD_RETRIES} attempts: {e}")
    raise last_error


def step3_download(status, headless=True, logger=None):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 3: Download")
    logger.log("=" * 60)

    confirmed = [r for r in status["records"]
                 if r["status"] == "confirmed" and not r.get("is_related")]
    unique_sns = list(dict.fromkeys(r["sn"] for r in confirmed))

    if not unique_sns:
        logger.log("沒有需要下載的 SN。")
        update_status_session(status, {"current_step": "download"})
        return status

    logger.log(f"需要下載：{len(unique_sns)} 個 SN")
    for sn in unique_sns:
        logger.log(f"  - {sn}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            args=["--ignore-certificate-errors"],
        )
        context = browser.new_context(
            ignore_https_errors=True,
            viewport={"width": 1920, "height": 1080},
        )
        page = context.new_page()

        try:
            for idx, sn in enumerate(unique_sns, 1):
                logger.log(f"\n{'─'*60}")
                logger.log(f"({idx}/{len(unique_sns)}) SN: {sn}")
                logger.log(f"{'─'*60}")

                # Get MAC
                mac = _get_mac_from_optools(page, sn, logger)
                if not mac:
                    error_msg = f"MAC address not found for SN {sn}"
                    logger.log(f"  [FAIL] {error_msg}")
                    for rec in status["records"]:
                        if rec["sn"] == sn and rec["status"] == "confirmed":
                            rec["status"] = "download_fail"
                            rec["error"] = error_msg
                    save_status(status)
                    continue

                # Get download link
                link = _download_from_optools(page, sn, mac, logger)
                if not link:
                    error_msg = "No download link found after polling"
                    logger.log(f"  [FAIL] {error_msg}")
                    for rec in status["records"]:
                        if rec["sn"] == sn and rec["status"] == "confirmed":
                            rec["status"] = "download_fail"
                            rec["error"] = error_msg
                    save_status(status)
                    continue

                # Download file — use first non-related record's temp_id as dir
                primary = next(
                    (r for r in status["records"]
                     if r["sn"] == sn and r["status"] == "confirmed"
                     and not r.get("is_related")),
                    None,
                )
                if not primary:
                    continue

                dest_dir = os.path.join(COREDUMPS_DIR, primary["temp_id"])
                filename = primary.get("coredump_files", "coredump")

                try:
                    dest_path = _download_file(link, dest_dir, filename, logger)

                    # Verify zip contents match coredump-key
                    ck = primary.get("coredump_key", "")
                    verified, _ = _verify_coredump(dest_path, ck, logger)

                    if verified:
                        for rec in status["records"]:
                            if rec["sn"] == sn and rec["status"] == "confirmed":
                                rec["status"] = "download_ok"
                                rec["download_path"] = dest_path
                        save_status(status)
                        logger.log(f"  [OK] SN {sn} 下載完成")
                    else:
                        error_msg = "OPTOOLS mis-match: coredump-key not found in zip"
                        for rec in status["records"]:
                            if rec["sn"] == sn and rec["status"] == "confirmed":
                                rec["status"] = "download_fail"
                                rec["error"] = error_msg
                        save_status(status)
                        logger.log(f"  [FAIL] SN {sn} 下載內容不符: {error_msg}")

                except Exception as e:
                    error_msg = str(e)
                    for rec in status["records"]:
                        if rec["sn"] == sn and rec["status"] == "confirmed":
                            rec["status"] = "download_fail"
                            rec["error"] = error_msg
                    save_status(status)
                    logger.log(f"  [FAIL] SN {sn} 下載失敗: {error_msg}")
        finally:
            context.close()
            browser.close()

    update_status_session(status, {"current_step": "download"})
    return status


# ═══════════════════════════════════════════════════════════════════════════════
# Step 4: Second Confirmation & Jira Post
# ═══════════════════════════════════════════════════════════════════════════════

def _build_adf_description(sn, model, daemon, firmware, jira_id):
    """Build Jira ADF description content."""
    year = datetime.now().strftime("%Y")
    ftp_url = f"ftp://{_config['ftp-host']}/jira/{year}_ML_Coredump_files/{jira_id}/"

    def _p(c): return {"type": "paragraph", "content": c}
    def _e(): return {"type": "paragraph", "content": []}
    def _b(t): return {"type": "text", "text": t, "marks": [{"type": "strong"}]}
    def _t(t): return {"type": "text", "text": t}
    def _l(u): return {"type": "text", "text": u, "marks": [{"type": "link", "attrs": {"href": u}}]}
    def _lt(u, t): return {"type": "text", "text": t, "marks": [{"type": "link", "attrs": {"href": u}}]}

    return [
        _p([_t("We have identified several coredump issues on customer devices "
               "through Elasticsearch monitoring. These incidents have been processed "
               "by the ELK Agent and automatically synchronized to Jira for tracking.")]),
        _e(),
        _p([_t("Please investigate and address these issues accordingly.")]),
        _e(),
        _p([_t("For your reference, all coredump instances identified from Elasticsearch "
               "have been compiled into the following SharePoint document:")]),
        _p([_t("\U0001f517 "), _lt(_config['sharepoint-url'], "Coredump Compilation List (SharePoint)")]),
        _e(),
        _p([_t("If you require any further clarification or technical details, "
               "please contact Max directly.")]),
        _e(),
        _p([_b("SN: "), _t(sn)]),
        _p([_b("Model: "), _t(model)]),
        _p([_b("Daemon: "), _t(daemon)]),
        _p([_b("Firmware: "), _t(firmware)]),
        _e(),
        _p([_t("You can download the coredump file from FTP site:")]),
        _p([_l(ftp_url)]),
    ]


def step4_post_jira(status, logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 4: 第二次確認 & Post Jira")
    logger.log("=" * 60)

    all_recs = status["records"]
    download_ok = [r for r in all_recs if r["status"] == "download_ok" and not r.get("is_related")]
    download_fail = [r for r in all_recs if r["status"] == "download_fail"]
    related = [r for r in all_recs if r.get("is_related") and r["status"] in ("confirmed", "download_ok")]
    need_confirm = download_ok

    # ── 4a: Summary ──
    logger.log(f"\n--- Summary ---")
    logger.log(f"  總筆數:         {len(all_recs)}")
    logger.log(f"  下載成功:       {len(download_ok)}")
    logger.log(f"  下載失敗:       {len(download_fail)}")
    logger.log(f"  重複 (related): {len(related)}")
    logger.log(f"  需要確認:       {len(need_confirm)}")

    if download_fail:
        logger.log(f"\n  下載失敗清單：")
        for r in download_fail:
            logger.log(f"    SN: {r['sn']}, Daemon: {r['daemon']}, Error: {r.get('error', 'unknown')}")

    # ── Mark download_fail records as OPTOOLS fail/mis-match in CSV ──
    if download_fail:
        csv_path, fieldnames, csv_rows = load_csv()
        for rec in download_fail:
            jira_col = rec.get("jira_col")
            temp_id = rec.get("temp_id")
            error = rec.get("error", "OPTOOLS fail")

            # Mark as "OPTOOLS fail" or more specific error message
            fail_mark = "OPTOOLS fail" if "OPTOOLS mis-match" not in error else "OPTOOLS mis-match"

            for r in csv_rows:
                if r.get(jira_col, "").strip() == temp_id:
                    r[jira_col] = fail_mark
            update_daily_csv(rec["_id"], jira_col, fail_mark)
            logger.log(f"  已標記 {jira_col}={fail_mark} (SN: {rec['sn']}, Error: {error})")
        save_csv(csv_path, fieldnames, csv_rows)
        save_status(status)

    if not need_confirm:
        logger.log("沒有需要 post 的項目。")
        update_status_session(status, {"current_step": "post"})
        return status

    # ── 4b: List ──
    logger.log(f"\n--- 待確認清單 ---")
    for idx, rec in enumerate(need_confirm, 1):
        logger.log(f"  ({idx}/{len(need_confirm)}) SN: {rec['sn']}, Daemon: {rec['daemon']}, "
                   f"Model: {rec['model']}, Key: {rec['coredump_key']}")
        logger.log(f"    File: {rec.get('download_path', 'N/A')}")

    # ── 4c: Per-record confirmation ──
    logger.log(f"\n--- 逐筆確認 ---")

    field_map = get_custom_field_map(_config, {
        'Coredump Daemon', 'Model', 'Coredump Information',
        'Function Category', 'Report Dept.', 'Severity',
        'Reproducible ?', 'Coredump File', 
    }, logger=logger)

    csv_path, fieldnames, csv_rows = load_csv()
    temp_to_jira = {}  # temp_id -> real jira_id

    for idx, rec in enumerate(need_confirm, 1):
        sn = rec["sn"]
        daemon = rec["daemon"]
        model = rec["model"]
        firmware = rec["firmware"]
        coredump_files = rec.get("coredump_files", "")
        # -- Max Add --
        fw = parse_affects_version(firmware) 
        logger.log(f"\n{'='*60}")
        logger.log(f"({idx}/{len(need_confirm)})")
        logger.log(f"  SN:               {sn}")
        logger.log(f"  Daemon:           {daemon}")
        logger.log(f"  Model:            {model}")
        logger.log(f"  Firmware:         {firmware}")
        logger.log(f"  fw:               {fw}")
        logger.log(f"  Coredump Key:     {rec['coredump_key']}")
        logger.log(f"  Coredump File(s): {coredump_files}")
        logger.log(f"  Download Path:    {rec.get('download_path', 'N/A')}")
        logger.log(f"  ITS:              {'Yes' if rec.get('is_its') else 'No'}")

        choice = ask_user(
            f"({idx}/{len(need_confirm)}) Post to Jira? (y=Post / s=跳過)",
            ["y", "s"], default="y",
        )

        if choice == "s":
            rec["status"] = "post_skipped"
            save_status(status)
            for r in csv_rows:
                if r.get(rec["jira_col"], "").strip() == rec["temp_id"]:
                    r[rec["jira_col"]] = "Skip"
            save_csv(csv_path, fieldnames, csv_rows)
            update_daily_csv(rec["_id"], rec["jira_col"], "Skip")
            logger.log(f"  已跳過")
            continue

        # ── Post to Jira ──
        summary_text = f"[AUTO][ML][coredump] {daemon} coredump (SN:{sn})"
        custom_fields = {}
        if 'Function Category' in field_map:
            custom_fields[field_map['Function Category']] = [{"value": "Coredump"}]
        if 'Report Dept.' in field_map:
            custom_fields[field_map['Report Dept.']] = {"value": "RD"}
        if 'Severity' in field_map:
            custom_fields[field_map['Severity']] = {"value": "L2"}
        if 'Reproducible ?' in field_map:
            custom_fields[field_map['Reproducible ?']] = {"value": "Yes"}
        if 'Coredump Information' in field_map:
            custom_fields[field_map['Coredump Information']] = coredump_files
        if 'Coredump File' in field_map:
            custom_fields[field_map['Coredump File']] = coredump_files
        custom_fields['customfield_10283'] = sn
        # -- Max Add --
        #custom_fields['customfield_10088'] = model

        # Create version if it doesn't exist (在 POST 之前檢查)
        try:
            # Get existing versions
            versions_url = f"{_config['JIRA_BASE_URL']}/rest/api/3/project/{PROJECT_KEY}/versions"
            versions_response = requests.get(
                versions_url,
                headers={"Accept": "application/json"},
                auth=HTTPBasicAuth(_config["EMAIL"], _config["API_TOKEN"])
            )
            existing_versions = []
            if versions_response.status_code == 200:
                existing_versions = [v.get('name', '') for v in versions_response.json()]

            # Create version if it doesn't exist
            if fw not in existing_versions:
                create_url = f"{_config['JIRA_BASE_URL']}/rest/api/3/version"
                create_response = requests.post(
                    create_url,
                    headers={
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                    },
                    auth=HTTPBasicAuth(_config["EMAIL"], _config["API_TOKEN"]),
                    data=json.dumps({
                        "name": fw,
                        "project": PROJECT_KEY,
                    })
                )
                if create_response.status_code in (200, 201):
                    logger.log(f"  已建立新版本: {fw}")
                else:
                    logger.log(f"  建立版本失敗 {fw}，HTTP {create_response.status_code}")
        except Exception as e:
            logger.log(f"  [WARN] 版本檢查/建立失敗: {e}")

        # 添加 versions 字段到 custom_fields (JIRA 標準欄位)
        custom_fields['versions'] = [{"name": fw}]

        try:
            issue = create_issue(
                _config, PROJECT_KEY, ISSUE_TYPE_NAME, summary_text,
                custom_fields=custom_fields, logger=logger,
            )
            jira_id = issue.get("key", "")

            # Model 和 versions 在 Create Screen 可能不支援，改用 PUT 更新
            put_fields = {}

            # 避免被Leo蓋掉, 等他做完
            time.sleep(9)

            # 更新 Model (自訂欄位)
            if 'Model' in field_map:
                put_fields[field_map['Model']] = [{"value": model}]
            else:
                logger.log(f"  [WARN] 找不到 'Model' 欄位，無法更新")

            # 更新 Affects Version (標準欄位)
            put_fields['versions'] = [{"name": fw}]

            # 只有在 put_fields 不是空的時候才更新
            if put_fields:
                try:
                    update_fields(_config, jira_id, put_fields, logger=logger)
                    logger.log(f"  已更新 Model 和 Affects Version")
                except Exception as e:
                    logger.log(f"  [WARN] 更新 Model/versions 失敗: {e}")

            # 驗證 Model 和 versions 是否已成功填入
            if 'Model' in field_map:
                try:
                    model_ok, versions_ok, model_val, versions_val = verify_issue_fields(
                        _config, jira_id, field_map['Model'], logger=logger
                    )
                    logger.log(f"  驗證結果:")
                    logger.log(f"    Model: {'✓ ' + model_val if model_ok else '✗ 未填入'}")
                    logger.log(f"    影響版本: {'✓ ' + versions_val if versions_ok else '✗ 未填入'}")
                    if not model_ok or not versions_ok:
                        logger.log(f"  [WARN] 欄位驗證失敗，請檢查 Jira issue {jira_id}")
                except Exception as e:
                    logger.log(f"  [WARN] 驗證欄位時發生錯誤: {e}")

            rec["jira_id"] = jira_id
            rec["status"] = "posted"
            temp_to_jira[rec["temp_id"]] = jira_id

            # Rename coredump directory: temp_id → jira_id
            old_dir = os.path.join(COREDUMPS_DIR, rec["temp_id"])
            new_dir = os.path.join(COREDUMPS_DIR, jira_id)
            if os.path.isdir(old_dir) and not os.path.exists(new_dir):
                os.rename(old_dir, new_dir)
                if rec.get("download_path"):
                    rec["download_path"] = rec["download_path"].replace(
                        rec["temp_id"], jira_id)
                logger.log(f"  Renamed: {rec['temp_id']} -> {jira_id}")

            # Update CSV
            for r in csv_rows:
                if r.get(rec["jira_col"], "").strip() == rec["temp_id"]:
                    r[rec["jira_col"]] = jira_id
            save_csv(csv_path, fieldnames, csv_rows)
            update_daily_csv(rec["_id"], rec["jira_col"], jira_id)

            # Update description
            try:
                adf = _build_adf_description(sn, model, daemon, firmware, jira_id)
                update_description(_config, jira_id, adf, logger=logger)
            except Exception as e:
                logger.log(f"  [WARN] 更新 description 失敗: {e}")

            save_status(status)
            logger.log(f"  [OK] Posted: {jira_id}")

        except Exception as e:
            logger.log(f"  [ERROR] Post 失敗: {e}")
            rec["status"] = "post_failed"
            rec["error"] = str(e)
            save_status(status)

    # ── Update related records with real jira_ids ──
    if temp_to_jira:
        csv_path, fieldnames, csv_rows = load_csv()
        for rec in related:
            real_jira_id = temp_to_jira.get(rec["temp_id"])
            if real_jira_id:
                rec["jira_id"] = real_jira_id
                rec["status"] = "posted"
                for r in csv_rows:
                    if r.get(rec["jira_col"], "").strip() == rec["temp_id"]:
                        r[rec["jira_col"]] = real_jira_id
                update_daily_csv(rec["_id"], rec["jira_col"], real_jira_id)
        save_csv(csv_path, fieldnames, csv_rows)
        save_status(status)
        logger.log(f"\n已更新 {len(temp_to_jira)} 個 related record 的 jira-id")

    update_status_session(status, {"current_step": "post"})
    return status


# ═══════════════════════════════════════════════════════════════════════════════
# Step 4b: Update jira-id count in ELK-summary
# ═══════════════════════════════════════════════════════════════════════════════

def update_jira_count(logger):
    """After all Jira posts, fill the 'count' column (after jira-id) in ELK-summary.csv
    and ELK-summary.xlsx.

    Logic: for each row whose jira-id is a valid ZNGA-xxxx, count how many rows
    in the entire CSV have that same value in 'related-jira-id', then write that
    count into the 'count' column positioned right after 'jira-id'.

    Uses raw csv.reader (index-based) to safely handle the duplicate 'count' header.
    If the second 'count' column does not yet exist, it is inserted automatically.
    After updating the CSV the same counts are written into ELK-summary.xlsx.
    """
    from collections import Counter

    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 4b: Update jira-id count")
    logger.log("=" * 60)

    csv_path = os.path.join(SCRIPT_DIR, "ELK-summary.csv")
    if not os.path.exists(csv_path):
        logger.log("  [SKIP] ELK-summary.csv not found.")
        return

    # ── Read raw rows ──
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        logger.log("  [SKIP] Empty CSV.")
        return

    header = all_rows[0]

    # Find column indices by name (handle duplicate 'count')
    jira_id_idx = None
    related_jira_id_idx = None
    new_count_idx = None   # second occurrence of 'count', positioned after jira-id
    count_seen = 0

    for i, col in enumerate(header):
        if col == "jira-id":
            jira_id_idx = i
        elif col == "related-jira-id":
            related_jira_id_idx = i
        elif col == "count":
            count_seen += 1
            if count_seen == 2:
                new_count_idx = i

    if jira_id_idx is None or related_jira_id_idx is None:
        logger.log(f"  [ERROR] Cannot find required columns. "
                   f"jira-id={jira_id_idx}, related-jira-id={related_jira_id_idx}")
        return

    # ── Insert second 'count' column after 'jira-id' if missing ──
    if new_count_idx is None:
        insert_at = jira_id_idx + 1
        header.insert(insert_at, "count")
        for row in all_rows[1:]:
            # Extend short rows before inserting
            while len(row) < insert_at:
                row.append("")
            row.insert(insert_at, "")
        new_count_idx = insert_at
        # Adjust related_jira_id_idx if it shifted right
        if related_jira_id_idx >= insert_at:
            related_jira_id_idx += 1
        logger.log(f"  已在 'jira-id' 後插入第二個 count 欄位 (index={new_count_idx})")

    data_rows = all_rows[1:]

    # ── Build related-jira-id counter ──
    related_counter = Counter()
    for row in data_rows:
        if len(row) > related_jira_id_idx:
            rid = row[related_jira_id_idx].strip()
            if re.match(r'^ZNGA-\d+$', rid):
                related_counter[rid] += 1

    # ── Fill new count column and collect summary ──
    summary = {}   # ZNGA-xxxx -> count written
    updated = 0

    for row in data_rows:
        if len(row) <= jira_id_idx:
            continue
        jira_id = row[jira_id_idx].strip()
        if not re.match(r'^ZNGA-\d+$', jira_id):
            continue

        cnt = related_counter.get(jira_id, 0)
        while len(row) <= new_count_idx:
            row.append("")
        row[new_count_idx] = str(cnt)
        summary[jira_id] = cnt
        updated += 1

    # ── Save CSV ──
    tmp_path = csv_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)
    os.replace(tmp_path, csv_path)

    # ── Also update XLSX (index-based, same logic as CSV) ──
    xlsx_path = _elk_query.SUMMARY_XLSX
    if os.path.exists(xlsx_path):
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb['in'] if 'in' in wb.sheetnames else wb.active

            # Find jira-id and second 'count' column in XLSX header
            xlsx_header = [ws.cell(row=1, column=c).value
                           for c in range(1, ws.max_column + 1)]
            xlsx_jira_id_idx = None
            xlsx_count_idx = None   # second 'count' (0-based)
            xlsx_count_seen = 0

            for i, col in enumerate(xlsx_header):
                if col == "jira-id":
                    xlsx_jira_id_idx = i
                elif col == "count":
                    xlsx_count_seen += 1
                    if xlsx_count_seen == 2:
                        xlsx_count_idx = i

            if xlsx_jira_id_idx is None:
                logger.log("  [WARN] XLSX 中找不到 jira-id 欄位，跳過 XLSX 更新")
            else:
                # Insert second 'count' column into XLSX if missing
                if xlsx_count_idx is None:
                    # openpyxl insert_cols uses 1-based index
                    insert_col_1based = xlsx_jira_id_idx + 2  # right after jira-id
                    ws.insert_cols(insert_col_1based)
                    ws.cell(row=1, column=insert_col_1based, value="count")
                    xlsx_count_idx = insert_col_1based - 1  # convert to 0-based
                    logger.log(f"  已在 XLSX 'jira-id' 後插入第二個 count 欄位")

                # Write count for every ZNGA-xxxx jira-id row
                xlsx_updated = 0
                for row_idx in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row_idx,
                                       column=xlsx_jira_id_idx + 1).value
                    jira_id_str = str(cell_val).strip() if cell_val else ""
                    if re.match(r'^ZNGA-\d+$', jira_id_str):
                        cnt = related_counter.get(jira_id_str, 0)
                        ws.cell(row=row_idx,
                                column=xlsx_count_idx + 1,
                                value=cnt)
                        xlsx_updated += 1

                wb.save(xlsx_path)
                logger.log(f"  已同步更新 XLSX {xlsx_updated} 筆 jira-id count 欄位")
        except Exception as e:
            logger.log(f"  [WARN] 更新 XLSX count 欄位失敗: {e}")

    # ── Summary log ──
    logger.log(f"\n  已更新 {updated} 筆 jira-id 的 count 欄位:")
    logger.log(f"  {'Jira ID':<15} {'count (related)':>15}")
    logger.log(f"  {'-'*15} {'-'*15}")
    for jid in sorted(summary):
        logger.log(f"  {jid:<15} {summary[jid]:>15}")
    logger.log(f"\n  related-jira-id 統計明細:")
    logger.log(f"  {'Jira ID':<15} {'related count':>15}")
    logger.log(f"  {'-'*15} {'-'*15}")
    for jid, cnt in sorted(related_counter.items()):
        logger.log(f"  {jid:<15} {cnt:>15}")


# ═══════════════════════════════════════════════════════════════════════════════
# Step 5: Upload FTP
# ═══════════════════════════════════════════════════════════════════════════════

def step5_upload_ftp(logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 5: Upload FTP")
    logger.log("=" * 60)

    if not os.path.isdir(COREDUMPS_DIR):
        logger.log("No coredumps directory. Skipping.")
        return

    # Skip tmp- directories (not yet posted to Jira)
    subfolders = [
        d for d in os.listdir(COREDUMPS_DIR)
        if os.path.isdir(os.path.join(COREDUMPS_DIR, d)) and not d.startswith("tmp-")
    ]
    if not subfolders:
        logger.log("No folders to upload (skipping tmp- directories).")
        return

    logger.log(f"Found {len(subfolders)} folder(s) to upload")

    year = datetime.now().strftime("%Y")
    remote_base = f"./jira/{year}_ML_Coredump_files"

    try:
        ftp = ftplib.FTP(_config["ftp-host"])
        ftp.login(_config["ftp-user"], _config["ftp-pass"])
        logger.log(f"[OK] FTP logged in as {_config['ftp-user']}")
    except ftplib.all_errors as e:
        logger.log(f"[ERROR] FTP connection failed: {e}")
        return

    try:
        # Ensure base directory
        try:
            ftp.cwd(remote_base)
            ftp.cwd("/")
        except ftplib.error_perm:
            ftp.mkd(remote_base)

        uploaded = 0
        skipped = 0

        for folder in subfolders:
            local_folder = os.path.join(COREDUMPS_DIR, folder)
            remote_folder = f"{remote_base}/{folder}"

            try:
                ftp.cwd(remote_folder)
                ftp.cwd("/")
            except ftplib.error_perm:
                ftp.mkd(remote_folder)

            try:
                existing = {os.path.basename(f) for f in ftp.nlst(remote_folder)}
            except ftplib.error_perm:
                existing = set()

            for filename in os.listdir(local_folder):
                local_file = os.path.join(local_folder, filename)
                if not os.path.isfile(local_file):
                    continue
                if filename in existing:
                    skipped += 1
                    continue
                remote_file = f"{remote_folder}/{filename}"
                with open(local_file, "rb") as f:
                    ftp.storbinary(f"STOR {remote_file}", f)
                uploaded += 1
                logger.log(f"  Uploaded: {remote_file}")

        logger.log(f"[DONE] Uploaded: {uploaded}, Skipped: {skipped}")
    except ftplib.all_errors as e:
        logger.log(f"[ERROR] FTP error: {e}")
    finally:
        try:
            ftp.quit()
        except ftplib.all_errors:
            ftp.close()


# ═══════════════════════════════════════════════════════════════════════════════
# Helper: Update XLSX from CSV (for markers like OPTOOLS fail)
# ═══════════════════════════════════════════════════════════════════════════════

def _update_xlsx_from_csv(logger):
    """Update existing records in XLSX from CSV (e.g., OPTOOLS fail markers).

    This ensures that when records fail download (marked OPTOOLS fail in CSV),
    those markers are reflected in the XLSX file even if the record already exists.
    """
    csv_path = _elk_query.SUMMARY_FILE
    xlsx_path = _elk_query.SUMMARY_XLSX

    if not os.path.exists(csv_path) or not os.path.exists(xlsx_path):
        return

    # Read CSV
    csv_rows = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            _id = row.get('_id', '').strip()
            if _id:
                csv_rows[_id] = row

    # Update XLSX
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb['in'] if 'in' in wb.sheetnames else wb.active

        # Get header mapping
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_map = {name: idx for idx, name in enumerate(header) if name}

        updated_count = 0
        # Update each row if it exists in CSV
        for row_idx in range(2, ws.max_row + 1):
            _id = ws.cell(row=row_idx, column=header_map.get('_id', 1) + 1).value
            if _id and _id in csv_rows:
                csv_row = csv_rows[_id]
                # Update each column from CSV
                for col_name, col_idx in header_map.items():
                    csv_val = csv_row.get(col_name, '')
                    ws.cell(row=row_idx, column=col_idx + 1, value=csv_val)
                updated_count += 1

        if updated_count > 0:
            wb.save(xlsx_path)
            logger.log(f"  Updated {updated_count} XLSX records from CSV")
    except Exception as e:
        logger.log(f"  Warning: Failed to update XLSX from CSV: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# Step 6: Merge XLSX
# ═══════════════════════════════════════════════════════════════════════════════

def step6_merge_xlsx(logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 6: Merge CSV → XLSX")
    logger.log("=" * 60)

    merge_logger = Logger(source="ELK-merge")
    _elk_query.merge_to_xlsx(_elk_query.SUMMARY_FILE, _elk_query.SUMMARY_XLSX, merge_logger)
    merge_logger.close()

    # ── Update existing XLSX records from CSV (e.g., OPTOOLS fail markers) ──
    _update_xlsx_from_csv(logger)

    logger.log("XLSX merge 完成")


# ═══════════════════════════════════════════════════════════════════════════════
# Step 7: Report
# ═══════════════════════════════════════════════════════════════════════════════

def step7_report(logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 7: Report")
    logger.log("=" * 60)

    _, _, rows = load_csv()

    jira_ids = {}  # jira_id -> [(col, sn, daemon)]
    for row in rows:
        for col in JIRA_ID_COLUMNS:
            jid = row.get(col, "").strip()
            if jid and jid not in SKIP_JIRA_VALUES and not jid.startswith("tmp-"):
                sn = row.get("sn", "").strip()
                daemon = row.get("daemon", "").strip()
                jira_ids.setdefault(jid, []).append((col, sn, daemon))

    unique_ids = sorted(jira_ids.keys())
    total = len(unique_ids)
    logger.log(f"有效 Jira ID: {total} 個")

    if total == 0:
        logger.log("沒有需要檢查的項目。")
        return

    year = datetime.now().strftime("%Y")
    remote_base = f"./jira/{year}_ML_Coredump_files"
    ftp_folders = {}
    ftp = None

    try:
        ftp = ftplib.FTP(_config["ftp-host"])
        ftp.login(_config["ftp-user"], _config["ftp-pass"])
        try:
            entries = ftp.nlst(remote_base)
            for entry in entries:
                name = os.path.basename(entry)
                path = f"{remote_base}/{name}"
                try:
                    files = ftp.nlst(path)
                    ftp_folders[name] = {
                        os.path.basename(f) for f in files
                        if os.path.basename(f) != name
                    }
                except ftplib.error_perm:
                    pass
        except ftplib.error_perm:
            pass
    except ftplib.all_errors as e:
        logger.log(f"[ERROR] FTP 連線失敗: {e}")

    jira_ok = jira_fail = ftp_ok = ftp_fail = 0
    problems = []

    for idx, jid in enumerate(unique_ids, 1):
        col, sn, daemon = jira_ids[jid][0]
        is_related = "related" in col

        exists, status_text = check_issue_exists(_config, jid)
        if exists:
            jira_ok += 1
            j_s = f"OK ({status_text})"
        else:
            jira_fail += 1
            j_s = "NOT FOUND"
            problems.append(f"  {jid}: Jira 不存在 (SN: {sn}, daemon: {daemon})")

        if is_related:
            f_s = "skip (related)"
        elif jid in ftp_folders and ftp_folders[jid]:
            ftp_ok += 1
            f_s = f"OK ({len(ftp_folders[jid])} file(s))"
        else:
            ftp_fail += 1
            f_s = "NOT FOUND" if jid not in ftp_folders else "EMPTY"
            problems.append(f"  {jid}: FTP 問題 (SN: {sn}, daemon: {daemon})")

        logger.log(f"  ({idx}/{total}) {jid}  Jira: {j_s}  FTP: {f_s}")

    if ftp:
        try:
            ftp.quit()
        except ftplib.all_errors:
            ftp.close()

    logger.log(f"\nReport: Jira {jira_ok} OK / {jira_fail} fail, "
               f"FTP {ftp_ok} OK / {ftp_fail} fail")
    if problems:
        logger.log(f"Problems ({len(problems)}):")
        for prob in problems:
            logger.log(prob)


# ═══════════════════════════════════════════════════════════════════════════════
# Step 8: Cleanup
# ═══════════════════════════════════════════════════════════════════════════════

def step8_cleanup(logger):
    logger.log("")
    logger.log("=" * 60)
    logger.log("Step 8: Cleanup")
    logger.log("=" * 60)

    # Clean logs
    if os.path.isdir(LOG_DIR):
        log_files = sorted(
            [f for f in os.listdir(LOG_DIR) if os.path.isfile(os.path.join(LOG_DIR, f))],
            key=lambda f: os.path.getmtime(os.path.join(LOG_DIR, f)),
        )
        if len(log_files) > MAX_LOG_FILES:
            to_delete = log_files[:len(log_files) - MAX_LOG_FILES]
            for f in to_delete:
                os.remove(os.path.join(LOG_DIR, f))
            logger.log(f"Logs: 刪除 {len(to_delete)} 個舊檔案（保留 {MAX_LOG_FILES} 個）")
        else:
            logger.log(f"Logs: {len(log_files)} 個檔案，未超過上限 {MAX_LOG_FILES}")

    # Clean coredumps
    if os.path.isdir(COREDUMPS_DIR):
        dirs = sorted(
            [d for d in os.listdir(COREDUMPS_DIR)
             if os.path.isdir(os.path.join(COREDUMPS_DIR, d))],
            key=lambda d: os.path.getmtime(os.path.join(COREDUMPS_DIR, d)),
        )
        if len(dirs) > MAX_COREDUMP_DIRS:
            to_delete = dirs[:len(dirs) - MAX_COREDUMP_DIRS]
            for d in to_delete:
                shutil.rmtree(os.path.join(COREDUMPS_DIR, d))
            logger.log(f"Coredumps: 刪除 {len(to_delete)} 個舊資料夾（保留 {MAX_COREDUMP_DIRS} 個）")
        else:
            logger.log(f"Coredumps: {len(dirs)} 個資料夾，未超過上限 {MAX_COREDUMP_DIRS}")


# ═══════════════════════════════════════════════════════════════════════════════
# Retry Mode
# ═══════════════════════════════════════════════════════════════════════════════

def run_retry(retry_sn=None, headless=True, logger=None):
    status = load_status()
    failed = [r for r in status.get("records", []) if r["status"] == "download_fail"]
    if retry_sn:
        failed = [r for r in failed if r["sn"] == retry_sn]

    if not failed:
        logger.log("沒有需要 retry 的項目。")
        return

    # Most recent first
    failed = list(reversed(failed))

    logger.log(f"失敗清單：{len(failed)} 筆")
    for idx, rec in enumerate(failed, 1):
        logger.log(f"  ({idx}) SN: {rec['sn']}, Error: {rec.get('error', 'unknown')}, "
                   f"Retries: {rec.get('retries', 0)}")

    for idx, rec in enumerate(failed, 1):
        choice = ask_user(
            f"({idx}/{len(failed)}) Retry SN {rec['sn']}? (y=重試 / s=跳過 / q=結束)",
            ["y", "s", "q"], default="y",
        )
        if choice == "q":
            break
        if choice == "y":
            rec["status"] = "confirmed"
            rec["error"] = None
            rec["retries"] = rec.get("retries", 0) + 1

    save_status(status)

    step3_download(status, headless=headless, logger=logger)
    step4_post_jira(status, logger)
    update_jira_count(logger)
    step5_upload_ftp(logger)
    step6_merge_xlsx(logger)
    step7_report(logger)


# ═══════════════════════════════════════════════════════════════════════════════
# Resume Mode
# ═══════════════════════════════════════════════════════════════════════════════

def run_resume(headless=True, logger=None):
    status = load_status()
    if not has_unfinished_work(status):
        logger.log("status.json 中沒有未完成的工作。")
        return None

    all_recs = status.get("records", [])
    confirmed = sum(1 for r in all_recs if r["status"] == "confirmed")
    download_ok = sum(1 for r in all_recs if r["status"] == "download_ok")
    download_fail = sum(1 for r in all_recs if r["status"] == "download_fail")
    posted = sum(1 for r in all_recs if r["status"] == "posted")
    elk_date = status.get("session", {}).get("elk_date", "unknown")

    logger.log(f"上次進度（{elk_date}）：")
    logger.log(f"  待下載:     {confirmed}")
    logger.log(f"  已下載:     {download_ok}")
    logger.log(f"  下載失敗:   {download_fail}")
    logger.log(f"  已完成:     {posted}")

    current_step = status.get("session", {}).get("current_step", "")

    if current_step in ("", "confirmed"):
        step3_download(status, headless=headless, logger=logger)
        step4_post_jira(status, logger)
        update_jira_count(logger)
        step5_upload_ftp(logger)
        step6_merge_xlsx(logger)
        step7_report(logger)
    elif current_step == "download":
        step3_download(status, headless=headless, logger=logger)
        step4_post_jira(status, logger)
        update_jira_count(logger)
        step5_upload_ftp(logger)
        step6_merge_xlsx(logger)
        step7_report(logger)
    elif current_step == "post":
        step4_post_jira(status, logger)
        update_jira_count(logger)
        step5_upload_ftp(logger)
        step6_merge_xlsx(logger)
        step7_report(logger)

    return elk_date


# ═══════════════════════════════════════════════════════════════════════════════
# Normal Pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def _run_normal(days=1, date=None, headless=True, logger=None):
    # Step 1
    elk_date = step1_elk_query(days=days, date=date, logger=logger)
    if not elk_date:
        return

    # Initialize status
    status = {
        "session": {
            "started": datetime.now().isoformat(),
            "updated": datetime.now().isoformat(),
            "elk_date": elk_date,
            "current_step": "query",
        },
        "records": [],
    }
    save_status(status)

    # Step 2
    status = step2_first_confirm(status, logger)
    confirmed = [r for r in status["records"] if r["status"] == "confirmed"]
    if not confirmed:
        logger.log("沒有確認的項目，跳過下載。")
        step6_merge_xlsx(logger)
        step7_report(logger)
        return

    # Step 3
    status = step3_download(status, headless=headless, logger=logger)

    # Step 4
    status = step4_post_jira(status, logger)

    # Step 4b
    update_jira_count(logger)

    # Step 5-7
    step5_upload_ftp(logger)
    step6_merge_xlsx(logger)
    step7_report(logger)


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Coredump Auto-Processing Tool v2")
    parser.add_argument("-d", "--days", type=int, default=1,
                        help="ELK query lookback days (default: 1)")
    parser.add_argument("--date", type=str, default=None,
                        help="Query specific date (YYYY-MM-DD)")
    parser.add_argument("--head", action="store_true", default=False,
                        help="Show browser (default: headless)")
    parser.add_argument("--resume", action="store_true", default=False,
                        help="Resume unfinished work")
    parser.add_argument("--retry", nargs="?", const="__all__", default=None,
                        metavar="SN", help="Retry failed downloads (optionally specify SN)")
    args = parser.parse_args()

    logger = Logger(source="coredump-v2")
    headless = not args.head

    try:
        # ── Retry mode ──
        if args.retry is not None:
            retry_sn = None if args.retry == "__all__" else args.retry
            run_retry(retry_sn=retry_sn, headless=headless, logger=logger)
            step8_cleanup(logger)
            logger.log("\nRetry 完成！")
            logger.close()
            return

        # ── Resume mode ──
        if args.resume:
            elk_date = run_resume(headless=headless, logger=logger)
            today = datetime.now(TZ_TAIPEI).strftime("%Y-%m-%d")
            if elk_date and elk_date != today:
                logger.log(f"\nResume 日期 ({elk_date}) 與今天 ({today}) 不同，"
                           "繼續執行今天的查詢...")
                _run_normal(days=args.days, date=args.date,
                            headless=headless, logger=logger)
            step8_cleanup(logger)
            logger.log("\nResume 完成！")
            logger.close()
            return

        # ── Normal mode: check for unfinished work ──
        status = load_status()
        if has_unfinished_work(status):
            elk_date = status.get("session", {}).get("elk_date", "unknown")
            all_recs = status.get("records", [])
            n_confirmed = sum(1 for r in all_recs if r["status"] == "confirmed")
            n_ok = sum(1 for r in all_recs if r["status"] == "download_ok")
            n_fail = sum(1 for r in all_recs if r["status"] == "download_fail")

            logger.log(f"發現上次未完成的工作（{elk_date}）：")
            logger.log(f"  待下載: {n_confirmed}, 已下載: {n_ok}, 下載失敗: {n_fail}")

            choice = ask_user(
                "(r=繼續上次進度 / n=重新開始 / q=離開)",
                ["r", "n", "q"], default="r",
            )
            if choice == "q":
                logger.log("使用者選擇離開。")
                logger.close()
                return
            if choice == "r":
                resume_date = run_resume(headless=headless, logger=logger)
                today = datetime.now(TZ_TAIPEI).strftime("%Y-%m-%d")
                if resume_date and resume_date != today:
                    logger.log(f"\nResume 日期 ({resume_date}) 與今天 ({today}) 不同，"
                               "繼續執行今天的查詢...")
                    _run_normal(days=args.days, date=args.date,
                                headless=headless, logger=logger)
                step8_cleanup(logger)
                logger.log("\n完成！")
                logger.close()
                return
            # choice == "n": fall through

        _run_normal(days=args.days, date=args.date, headless=headless, logger=logger)
        step8_cleanup(logger)
        logger.log("\n完成！")

    except KeyboardInterrupt:
        logger.log("\n使用者中斷 (Ctrl+C)")
    except Exception as e:
        logger.log(f"\n[ERROR] {e}")
        import traceback
        logger.log(traceback.format_exc())
    finally:
        logger.close()


if __name__ == "__main__":
    main()
