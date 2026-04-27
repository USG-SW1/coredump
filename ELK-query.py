#!/usr/bin/env python3
"""
Query Elasticsearch for infection index data and export to CSV.
- Time range: yesterday (full day, Asia/Taipei timezone)
- Filters: firmware contains '1.38', target contains '.core.zip'
- Output: dated CSV + merge into ELK-summary.csv
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from elasticsearch import Elasticsearch
import openpyxl

from config_loader import load_config
from logger import Logger

_config = load_config(extra_keys=["ES-url"])
ES_URL = _config["ES-url"]
INDEX = "infection"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
SUMMARY_FILE = os.path.join(OUTPUT_DIR, "ELK-summary.csv")
SUMMARY_XLSX = os.path.join(OUTPUT_DIR, "ELK-summary.xlsx")

TZ_TAIPEI = timezone(timedelta(hours=8))


# All known fields from ES mapping, in fixed order
KNOWN_FIELDS = [
    "_id", "time", "sn", "model", "firmware", "fw_major_version",
    "major_version", "major_verion", "daemon", "category", "target",
    "coredump", "action", "mode", "type", "count", "status",
    "apply_rule", "infection_status", "message", "message_normalized",
    "jira-id", "related-jira-id",
    "ITS-jira-id", "ITS-related-jira-id",
]


def get_date_range(days=1, specific_date=None):
    """Return date range in ISO 8601 (UTC+8).

    If specific_date is given (YYYY-MM-DD), return that single day's range.
    Otherwise, return from (today - days) to yesterday.
    """
    if specific_date:
        day = datetime.strptime(specific_date, "%Y-%m-%d").replace(tzinfo=TZ_TAIPEI)
        start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
        return start.isoformat(), end.isoformat()
    now = datetime.now(TZ_TAIPEI)
    start_day = now - timedelta(days=days)
    end_day = now - timedelta(days=1)
    start = start_day.replace(hour=0, minute=0, second=0, microsecond=0)
    end = end_day.replace(hour=23, minute=59, second=59, microsecond=999999)
    return start.isoformat(), end.isoformat()


def _create_es_client():
    """Create Elasticsearch client with retry and timeout settings."""
    return Elasticsearch(
        ES_URL,
        request_timeout=60,
        max_retries=3,
        retry_on_timeout=True,
    )


def query_elasticsearch(start_time, end_time, logger=None):
    """Query ES with filters and return all hits using search_after pagination."""
    es = _create_es_client()

    query = {
        "bool": {
            "must": [
                {
                    "range": {
                        "time": {
                            "gte": start_time,
                            "lte": end_time
                        }
                    }
                },
                {
                    "wildcard": {
                        "firmware": {
                            "value": "*1.38*"
                        }
                    }
                },
                {
                    "wildcard": {
                        "target": {
                            "value": "*.core.zip*"
                        }
                    }
                }
            ]
        }
    }
    sort = [{"time": "asc"}, {"_doc": "asc"}]

    # First request
    body = {"size": 10000, "query": query, "sort": sort}
    resp = es.search(index=INDEX, body=body)

    hits = resp.get("hits", {}).get("hits", [])
    all_sources = [{**hit["_source"], "_id": hit["_id"]} for hit in hits]

    total = resp.get("hits", {}).get("total", {})
    if isinstance(total, dict):
        total_count = total.get("value", len(hits))
    else:
        total_count = total

    msg = f"Total matched: {total_count}, fetched so far: {len(all_sources)}"
    logger.log(msg)

    # Paginate with search_after
    while len(hits) > 0:
        last_sort = hits[-1].get("sort")
        if not last_sort:
            break
        body["search_after"] = last_sort
        resp = es.search(index=INDEX, body=body)
        hits = resp.get("hits", {}).get("hits", [])
        all_sources.extend({**hit["_source"], "_id": hit["_id"]} for hit in hits)
        if hits:
            msg = f"  Paginated: fetched {len(all_sources)} so far..."
            logger.log(msg)

    msg = f"Total fetched: {len(all_sources)}"
    logger.log(msg)
    logger.log_records(all_sources, label="ELK raw query data")
    return all_sources


def write_csv(records, filepath, logger=None):
    """Write records to CSV with all known fields in fixed order."""
    if not records:
        if logger:
            logger.log("No records to write.")
        return

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=KNOWN_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)

    if logger:
        logger.log(f"Written {len(records)} records to {filepath}")


def _target_dedup_key(target):
    """Extract dedup key from target path.

    Strip timestamp prefix (e.g. '260227-114818-') and firmware version prefix
    (e.g. '1.37_ABZI.1_') from filename before comparison.
    Example: '260227-114818-1.37_ABZI.1_-45c_libfp-ips-fp-rte:2.core.zip'
          -> '-45c_libfp-ips-fp-rte:2.core.zip'
    """
    if not target:
        return ""
    filename = os.path.basename(target)
    # Skip first 14 chars (timestamp portion like '260227-114818-')
    suffix = filename[14:] if len(filename) > 14 else filename
    # Strip firmware version prefix (e.g. '1.37_ABZI.1_')
    suffix = re.sub(r'^\d+\.\d+_[A-Za-z]+\.\d+_', '', suffix)
    return suffix


def _make_dedup_key(row):
    """Create a dedup key from target_suffix only."""
    return _target_dedup_key(row.get("target", ""))




def _backfill_jira_to_daily(daily_csv, backfill_updates, merge_logger):
    """Backfill jira fields from summary to daily CSV.

    backfill_updates: {_id: {col: value, ...}, ...}
    """
    with open(daily_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    updated_count = 0
    for row in rows:
        rid = row.get("_id", "")
        if rid in backfill_updates:
            for col, val in backfill_updates[rid].items():
                row[col] = val
            updated_count += 1
            merge_logger.log(f"  BACKFILL: _id={rid}, jira={backfill_updates[rid]}")

    if updated_count:
        with open(daily_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        merge_logger.log(f"Backfilled jira fields to {updated_count} rows in {daily_csv}")


def merge_to_summary(daily_csv, summary_file, merge_logger):
    """Append daily CSV data to summary file. If summary doesn't exist, copy daily as summary.
    Every record is logged to merge_logger."""
    if not os.path.exists(daily_csv):
        merge_logger.log(f"Daily CSV {daily_csv} not found, skipping merge.")
        return

    if not os.path.exists(summary_file):
        # First time: read daily CSV and dedup before writing as summary
        with open(daily_csv, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            header = list(reader.fieldnames)
            daily_rows = list(reader)
        seen_keys = set()
        deduped_rows = []
        for row in daily_rows:
            dk = _make_dedup_key(row)
            if dk not in seen_keys:
                seen_keys.add(dk)
                deduped_rows.append(row)
                merge_logger.log_record("ADDED", row)
            else:
                merge_logger.log_record("SKIPPED_DUP", row)
        skipped = len(daily_rows) - len(deduped_rows)
        with open(summary_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=header)
            writer.writeheader()
            for row in deduped_rows:
                writer.writerow(row)
        if skipped:
            merge_logger.log(f"Skipped {skipped} duplicate rows (by target_suffix)")
        merge_logger.log(f"Created {summary_file} with {len(deduped_rows)} rows from {daily_csv}")
    else:
        # Read existing summary to collect existing _id set
        with open(summary_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            summary_header = list(reader.fieldnames)
            existing_rows = list(reader)
        existing_ids = {row.get("_id") for row in existing_rows if row.get("_id")}
        existing_dedup_map = {}
        for row in existing_rows:
            dk = _make_dedup_key(row)
            if dk and dk not in existing_dedup_map:
                existing_dedup_map[dk] = row

        # Read daily rows as dicts
        with open(daily_csv, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            daily_header = list(reader.fieldnames)
            daily_rows = list(reader)

        # Filter out duplicates: by _id or by target_suffix
        jira_cols = ["jira-id", "related-jira-id", "ITS-jira-id", "ITS-related-jira-id"]
        new_rows = []
        backfill_updates = {}  # _id -> {col: value}
        for row in daily_rows:
            if row.get("_id") in existing_ids:
                merge_logger.log_record("SKIPPED_DUP_ID", row)
                continue
            dk = _make_dedup_key(row)
            if dk in existing_dedup_map:
                merge_logger.log_record("SKIPPED_DUP_KEY", row)
                # Collect jira fields from existing summary row for backfill
                existing_row = existing_dedup_map[dk]
                jira_updates = {}
                for col in jira_cols:
                    val = existing_row.get(col, "").strip()
                    if val:
                        jira_updates[col] = val
                if jira_updates:
                    backfill_updates[row.get("_id", "")] = jira_updates
                continue
            new_rows.append(row)
            existing_dedup_map[dk] = row
            merge_logger.log_record("ADDED", row)
        skipped = len(daily_rows) - len(new_rows)
        if skipped:
            merge_logger.log(f"Skipped {skipped} duplicate rows (by _id or target_suffix)")

        # Backfill jira fields to daily CSV
        if backfill_updates:
            _backfill_jira_to_daily(daily_csv, backfill_updates, merge_logger)

        if not new_rows:
            merge_logger.log("No new rows to merge to CSV.")
            return

        # Merge headers if needed
        merged_header = list(summary_header)
        for col in daily_header:
            if col not in merged_header:
                merged_header.append(col)

        if merged_header != summary_header:
            # Headers changed, rewrite entire file
            with open(summary_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=merged_header)
                writer.writeheader()
                for row in existing_rows:
                    writer.writerow(row)
                for row in new_rows:
                    writer.writerow(row)
            merge_logger.log(f"Merged {len(new_rows)} rows into {summary_file} (headers updated)")
        else:
            # Same headers, just append
            with open(summary_file, "a", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=summary_header)
                for row in new_rows:
                    writer.writerow(row)
            merge_logger.log(f"Appended {len(new_rows)} rows to {summary_file}")


def merge_to_xlsx(summary_csv, xlsx_file, merge_logger):
    """Merge ELK-summary.csv into ELK-summary.xlsx (sheet 'in').
    Check duplicates by _id and dedup key before adding. Log every record."""
    if not os.path.exists(summary_csv):
        merge_logger.log(f"CSV {summary_csv} not found, skipping xlsx merge.")
        return

    # Read CSV rows
    with open(summary_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        csv_header = list(reader.fieldnames)
        csv_rows = list(reader)

    merge_logger.log(f"=== Merge CSV -> XLSX: {len(csv_rows)} CSV rows ===")

    sheet_name = "in"

    if os.path.exists(xlsx_file):
        wb = openpyxl.load_workbook(xlsx_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            # Write header
            for col_idx, col_name in enumerate(csv_header, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for col_idx, col_name in enumerate(csv_header, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

    # Read existing xlsx header and data
    xlsx_header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Build header-to-column index mapping
    header_col = {name: idx for idx, name in enumerate(xlsx_header) if name}

    # Collect existing _ids and dedup keys from xlsx
    existing_ids = set()
    existing_dedup_keys = set()
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_name, col_idx in header_col.items():
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            row_data[col_name] = str(val) if val is not None else ""
        _id = row_data.get("_id", "")
        if _id:
            existing_ids.add(_id)
        existing_dedup_keys.add(_make_dedup_key(row_data))

    merge_logger.log(f"XLSX existing rows: {ws.max_row - 1}, existing _ids: {len(existing_ids)}")

    # Check for new header columns and add if needed
    for col_name in csv_header:
        if col_name not in header_col:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            header_col[col_name] = new_col - 1
            xlsx_header.append(col_name)

    # Merge new rows
    added = 0
    skipped = 0
    for row in csv_rows:
        _id = row.get("_id", "")
        if _id in existing_ids:
            merge_logger.log_record("XLSX_SKIPPED_DUP_ID", row)
            skipped += 1
            continue
        dk = _make_dedup_key(row)
        if dk in existing_dedup_keys:
            merge_logger.log_record("XLSX_SKIPPED_DUP_KEY", row)
            skipped += 1
            continue

        # Append new row
        new_row_idx = ws.max_row + 1
        for col_name in csv_header:
            col_idx = header_col.get(col_name)
            if col_idx is not None:
                ws.cell(row=new_row_idx, column=col_idx + 1, value=row.get(col_name, ""))
        existing_ids.add(_id)
        existing_dedup_keys.add(dk)
        merge_logger.log_record("XLSX_ADDED", row)
        added += 1

    wb.save(xlsx_file)
    merge_logger.log(f"XLSX merge done: {added} added, {skipped} skipped (duplicates). Total rows: {ws.max_row - 1}")


def main():
    parser = argparse.ArgumentParser(description="Query ELK infection index and export to CSV")
    parser.add_argument("-d", "--days", type=int, default=1,
                        help="Number of days to look back (default: 1 = yesterday only)")
    parser.add_argument("--date", type=str, default=None,
                        help="Specific date to query (YYYY-MM-DD), overrides --days")
    parser.add_argument("--merge-xlsx", action="store_true",
                        help="Only merge ELK-summary.csv into ELK-summary.xlsx (skip ELK query)")
    args = parser.parse_args()

    # --merge-xlsx mode: only merge CSV to XLSX, then exit
    if args.merge_xlsx:
        merge_logger = Logger(source="ELK-merge")
        merge_logger.log("=== XLSX merge-only mode ===")
        merge_to_xlsx(SUMMARY_FILE, SUMMARY_XLSX, merge_logger)
        merge_logger.log("=== XLSX merge-only done ===")
        merge_logger.close()
        return

    logger = Logger()

    start_time, end_time = get_date_range(args.days, specific_date=args.date)
    logger.log(f"Querying index '{INDEX}' for time range: {start_time} ~ {end_time}")
    logger.log(f"Filters: firmware=*1.38*, target=*.core.zip*")

    records = query_elasticsearch(start_time, end_time, logger=logger)

    if not records:
        logger.log("No matching records found. Exiting.")
        logger.close()
        sys.exit(0)

    # Daily CSV filename: use queried date, not today's date
    if args.date:
        date_str = args.date
    else:
        # --days mode: use the end date (yesterday) of the query range
        end_day = datetime.now(TZ_TAIPEI) - timedelta(days=1)
        date_str = end_day.strftime("%Y-%m-%d")
    daily_csv = os.path.join(OUTPUT_DIR, f"{date_str}.csv")

    write_csv(records, daily_csv, logger=logger)
    logger.log(f"Daily CSV written: {daily_csv}")

    merge_logger = Logger(source="ELK-merge")
    merge_logger.log(f"=== Merge session started: {daily_csv} ===")

    merge_to_summary(daily_csv, SUMMARY_FILE, merge_logger)

    merge_logger.log(f"=== Merge session ended ===")
    merge_logger.close()

    logger.log("Done.")
    logger.close()


if __name__ == "__main__":
    main()
